import pandas as pd
from db_classes import *
from openpyxl import Workbook
from openpyxl import load_workbook
import random, math
from datetime import datetime

class ExcelLogger:
    SHEET_NAME = "Log Datasheet"

    def __init__(self, file):
        self.__file = file

    def ReadFile(self):
        df = pd.read_excel(self.__file, sheet_name = self.SHEET_NAME)
        return df

    def ParseFileAndCheckDuplicated(self):
        try:
            workbook = load_workbook(filename=self.__file)
        except Exception as e:
            if(isinstance(e,FileNotFoundError)):
                print("File not existed, new one will be created")
                workbook = self.__CreateNew()
            else:
                print(f'{e} ------> exitting')
                return

        datadict = {}
        sheet = self.__GetLogDataSheet(workbook)
        isKeyDuplicated = False
        for row in sheet.values:
            if(row[0] == TEXT_1ST_ROW_1ST_COL):
                continue

            if(row[IWCODE_IDX] in datadict):
                isKeyDuplicated = True
                #TODO: log error

            datadict[row[IWCODE_IDX]] = row[DATA_IDX]
            col = DATA_IDX + 1
            while(col < len(row) and row[col] != None):
                datadict[row[IWCODE_IDX]] += f',{row[col]}'
                col += 1

        return datadict, isKeyDuplicated, workbook

    def Log(self, wb, code, user, datalist):
        time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws = self.__GetLogDataSheet(wb)
        FoundExisting = False
        for i in range(0, len(datalist)):
            if code == datalist[i][IWCODE_IDX]:
                row, userCol = self.__FindBlankCellByRow(ws, i+2)
                FoundExisting = True
                break

        if(FoundExisting):
            timeCol = userCol + 1 
        else:
            row = len(datalist) + 2
            codeCol = IWCODE_COL
            ws.cell(row,codeCol).value = code

            userCol = USER_COL
            timeCol = TIME_COL

        ws.cell(row,userCol).value = user
        ws.cell(row,timeCol).value = time
        print(f'Wrote {user} to [{row}][{userCol}] and {time} to [{row}][{timeCol}]')
        wb.save(self.__file)

    def __GetLogDataSheet(self, wb):
        try:
            ws = wb[self.SHEET_NAME]
        except KeyError:
            print(f'No existed sheet. Creating new one as {self.SHEET_NAME}')
            ws = wb.create_sheet(self.SHEET_NAME)
            ws.append([TEXT_1ST_ROW_1ST_COL, TEXT_1ST_ROW_2ND_COL])
        finally:
            ws = wb[self.SHEET_NAME]
    
        wb.save(self.__file)
        return ws

    def __FindBlankCellByRow(self, ws ,row):
        col = 1
        while(ws.cell(row = row, column = col).value != None):
            col = col + 1

        return row, col
    
    def __CreateNew(self):
        workbook = Workbook()
        sheet = self.__GetLogDataSheet(workbook)
        workbook.save(self.__file)
        return workbook
